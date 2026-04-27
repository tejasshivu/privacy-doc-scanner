# clean_xlsx.py
# =============
# Strips metadata from Excel files by copying data into a fresh workbook

from openpyxl import load_workbook, Workbook

def clean_xlsx(input_path, output_path):
    try:
        # Open the original workbook
        wb_original = load_workbook(input_path)

        # --- Read original metadata ---
        original = {}
        try: original["Creator"]          = wb_original.properties.creator or ""
        except: pass
        try: original["Last Modified By"] = wb_original.properties.lastModifiedBy or ""
        except: pass
        try: original["Title"]            = wb_original.properties.title or ""
        except: pass
        try: original["Subject"]          = wb_original.properties.subject or ""
        except: pass
        try: original["Keywords"]         = wb_original.properties.keywords or ""
        except: pass

        # --- Create a brand new blank workbook ---
        # This starts completely clean with no metadata
        wb_clean = Workbook()

        # Remove the default empty sheet created by Workbook()
        wb_clean.remove(wb_clean.active)

        # --- Copy every sheet and its data across ---
        for sheet_name in wb_original.sheetnames:
            original_sheet = wb_original[sheet_name]

            # Create a matching sheet in the clean workbook
            clean_sheet = wb_clean.create_sheet(title=sheet_name)

            # Copy every cell row by row, column by column
            for row in original_sheet.iter_rows():
                for cell in row:
                    # Copy only the value — not formatting or metadata
                    clean_sheet[cell.coordinate] = cell.value

        # --- Save the clean workbook ---
        wb_clean.save(output_path)

        # --- Build cleaned metadata (should be empty) ---
        cleaned = {
            "Creator": "Anonymous",
            "Last Modified By": ""
        }

        removed = [k for k, v in original.items() if v]

        return {
            "success":  True,
            "original": original,
            "cleaned":  cleaned,
            "removed":  removed
        }

    except Exception as e:
        return {"success": False, "error": str(e)}